import streamlit as st
from openpyxl import load_workbook
from datetime import datetime
import time

st.set_page_config(page_title="Brandon Finances Pro", page_icon="⚡", layout="wide")

archivo = 'REGISTRO INGAS 2026 TEMP 9 - copia.xlsx'
meses_dict = {1:"ENE", 2:"FEB", 3:"MAR", 4:"ABR", 5:"MAY", 6:"JUN", 
              7:"JUL", 8:"AGO", 9:"SEP", 10:"OCT", 11:"NOV", 12:"DIC"}

SUPERMERCADOS = ["MERCADONA", "CARREFOUR", "LIDL", "ALDI", "DIA", "CONSUM", "ALCAMPO", "AHORRAMAS"]

def safe_float(val):
    try: return float(val)
    except: return 0.0

@st.cache_data(ttl=600)
def leer_datos_excel():
    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb[meses_dict[datetime.now().month]]
        
        def leer_celda(celda, def_val):
            return str(ws[celda].value).upper().strip() if ws[celda].value else def_val
        
        fijos = {
            "ALQUILER": {"concepto": leer_celda('C12', "ALQUILER"), "categoria": "ALQUILER", "fila": 64},
            "LUZ":      {"concepto": leer_celda('C13', "LUZ"), "categoria": "LUZ", "fila": 67},
            "AGUA":     {"concepto": leer_celda('C14', "AGUA"), "categoria": "AGUA", "fila": 68},
            "GAS":      {"concepto": leer_celda('C15', "GAS"), "categoria": "GAS", "fila": 69},
            "INTERNET": {"concepto": "INTERNET M. PELAYO (20€/3)", "categoria": "INTERNET/LÍNEA MÓVIL", "fila": 70},
            "MOVIL":    {"concepto": "LINEA MOVIL LOWI", "categoria": "INTERNET/LÍNEA MÓVIL", "fila": 71}
        }
        
        variables = []
        for i in range(18, 54):
            val = ws[f'C{i}'].value
            if val:
                cat = str(val).upper().strip()
                if cat not in variables:
                    variables.append(cat)
        
        if not variables:
            variables = ["OTROS GASTOS VARIABLES"]
            
        # --- MEJORA 4: LECTURA DEL PANEL DE CONTROL ---
        dashboard = {
            "comida": safe_float(ws['J17'].value),
            "variables": safe_float(ws['L18'].value),
            "total": safe_float(ws['J54'].value)
        }
            
        return fijos, variables, dashboard
        
    except:
        fijos = {
            "ALQUILER": {"concepto": "ALQUILER", "categoria": "ALQUILER", "fila": 64},
            "LUZ":      {"concepto": "LUZ", "categoria": "LUZ", "fila": 67},
            "AGUA":     {"concepto": "AGUA", "categoria": "AGUA", "fila": 68},
            "GAS":      {"concepto": "GAS", "categoria": "GAS", "fila": 69},
            "INTERNET": {"concepto": "INTERNET M. PELAYO (20€/3)", "categoria": "INTERNET/LÍNEA MÓVIL", "fila": 70},
            "MOVIL":    {"concepto": "LINEA MOVIL LOWI", "categoria": "INTERNET/LÍNEA MÓVIL", "fila": 71}
        }
        return fijos, ["OTROS GASTOS VARIABLES"], {"comida": 0.0, "variables": 0.0, "total": 0.0}

mapa_fijos, categorias_variables, dashboard = leer_datos_excel()

cat_suministros = ["LUZ", "AGUA", "GAS", "INTERNET/LÍNEA MÓVIL"]
conceptos_suministros = [mapa_fijos["LUZ"]["concepto"], mapa_fijos["AGUA"]["concepto"], mapa_fijos["GAS"]["concepto"], mapa_fijos["INTERNET"]["concepto"], mapa_fijos["MOVIL"]["concepto"]]

conceptos_conocidos = [mapa_fijos["ALQUILER"]["concepto"]] + conceptos_suministros + SUPERMERCADOS
opciones_desplegable = ["✍️ NUEVO GASTO..."] + conceptos_conocidos

# --- MEJORA 3: BÚSQUEDA SEGURA DE FILAS ---
def encontrar_fila_segura(ws, inicio, fin):
    for r in range(fin, inicio - 1, -1):
        if ws.cell(row=r, column=3).value is not None:
            if r < fin: return r + 1
            else: raise ValueError(f"La tabla está llena (límite fila {fin}).")
    return inicio

if 'paso' not in st.session_state: st.session_state.paso = "formulario"

def finalizar_y_limpiar():
    st.cache_data.clear()
    for k in ['concepto_temp', 'importe_temp', 'detalles_temp', 'fecha_temp', 'cat_temp']:
        if k in st.session_state:
            del st.session_state[k]
    st.session_state.paso = "formulario"
    st.rerun()

st.title("🏦 Panel de Registro - Brandon")

# --- PANEL DE CONTROL (DASHBOARD) ---
col_d1, col_d2, col_d3 = st.columns(3)
col_d1.metric("🛒 Comida Mes (J17)", f"{dashboard['comida']:.2f} €")
col_d2.metric("💳 Variables Mes (L18)", f"{dashboard['variables']:.2f} €")
col_d3.metric("🔥 TOTAL MES (J54)", f"{dashboard['total']:.2f} €")
st.divider()

# --- MEJORA 2: BOTÓN DEL PÁNICO (DESHACER) ---
if 'ultimo_guardado' in st.session_state:
    st.info(f"✅ Último registro: **{st.session_state.ultimo_guardado['concepto']}** ({st.session_state.ultimo_guardado['importe']}€) en fila {st.session_state.ultimo_guardado['fila']}")
    if st.button("🗑️ Deshacer / Borrar último registro", type="primary"):
        try:
            wb = load_workbook(archivo)
            ws = wb[st.session_state.ultimo_guardado['mes']]
            fila_borrar = st.session_state.ultimo_guardado['fila']

            ws.cell(row=fila_borrar, column=3).value = None
            ws.cell(row=fila_borrar, column=4).value = 0.0
            ws.cell(row=fila_borrar, column=5).value = None
            ws.cell(row=fila_borrar, column=6).value = None
            ws.cell(row=fila_borrar, column=7).value = None

            wb.save(archivo)
            del st.session_state.ultimo_guardado
            st.cache_data.clear()
            st.success("Registro eliminado. Cantidad fijada a 0.00€.")
            time.sleep(1)
            st.rerun()
        except Exception as e:
            st.error(f"Error al borrar: Cierra el Excel. ({e})")
    st.divider()

if st.session_state.paso == "formulario":
    with st.container(border=True):
        col1, col2 = st.columns(2)
        
        with col1:
            concepto_sel = st.selectbox("Concepto:", opciones_desplegable)
            
            if concepto_sel == "✍️ NUEVO GASTO...":
                concepto_real = st.text_input("Escribe el nombre del gasto:")
                opciones_cat = categorias_variables 
                cat_predeterminada = opciones_cat[0] if opciones_cat else "OTROS"
            else:
                concepto_real = concepto_sel
                con_up = concepto_real.upper() if concepto_real else ""
                
                if con_up == mapa_fijos["ALQUILER"]["concepto"] or "ALQU" in con_up:
                    opciones_cat = ["ALQUILER"]
                elif con_up in conceptos_suministros or any(con_up in s for s in conceptos_suministros):
                    opciones_cat = cat_suministros
                elif con_up in SUPERMERCADOS or "COMIDA" in con_up:
                    opciones_cat = ["COMIDA"]
                else:
                    opciones_cat = ["OTROS"]
                
                cat_predeterminada = opciones_cat[0]
                for datos in mapa_fijos.values():
                    if datos["concepto"] == con_up:
                        cat_predeterminada = datos["categoria"]
                        break

            idx = opciones_cat.index(cat_predeterminada) if cat_predeterminada in opciones_cat else 0
            categoria_sel = st.selectbox("Categoría oficial:", opciones_cat, index=idx)
            
        with col2:
            importe = st.number_input("Cantidad (€)", min_value=0.0, format="%.2f")
            fecha = st.date_input("Fecha", datetime.now())
            detalles = st.text_area("Detalles / Comentarios")

        if st.button("🚀 REGISTRAR GASTO", use_container_width=True):
            if concepto_real and importe > 0:
                st.session_state.concepto_temp = concepto_real.upper()
                st.session_state.importe_temp = importe
                st.session_state.fecha_temp = fecha
                st.session_state.detalles_temp = detalles
                st.session_state.cat_temp = categoria_sel
                st.session_state.paso = "confirmacion"
                st.rerun()
            else:
                st.warning("Escribe el concepto y el importe mayor a 0.")

elif st.session_state.paso == "confirmacion":
    st.subheader("⚠️ Revisa los datos antes de guardar:")
    
    with st.container(border=True):
        st.markdown(f"📝 **Concepto (Col C):** {st.session_state.concepto_temp}")
        st.markdown(f"💶 **Cantidad (Col D):** {st.session_state.importe_temp:.2f} €")
        st.markdown(f"📅 **Fecha (Col E):** {st.session_state.fecha_temp.strftime('%d/%m/%Y')}")
        st.markdown(f"📂 **Categoría (Col F):** {st.session_state.cat_temp}")
        st.markdown(f"💬 **Detalles (Col G):** {st.session_state.detalles_temp.upper() if st.session_state.detalles_temp else '-'}")
    
    c_si, c_no = st.columns(2)
    
    if c_si.button("✅ SÍ, TODO CORRECTO", use_container_width=True):
        try:
            wb = load_workbook(archivo)
            mes_p = meses_dict[st.session_state.fecha_temp.month]
            ws = wb[mes_p]
            
            fila_destino = 108 
            es_fijo = False
            
            for datos in mapa_fijos.values():
                if st.session_state.concepto_temp == datos["concepto"]:
                    fila_destino = datos["fila"]
                    es_fijo = True
                    break
            
            if not es_fijo:
                if st.session_state.cat_temp == "COMIDA":
                    fila_destino = encontrar_fila_segura(ws, 74, 105)
                else:
                    fila_destino = encontrar_fila_segura(ws, 108, 160)

            ws.cell(row=fila_destino, column=3).value = st.session_state.concepto_temp
            ws.cell(row=fila_destino, column=4).value = st.session_state.importe_temp
            ws.cell(row=fila_destino, column=5).value = st.session_state.fecha_temp.strftime("%d/%m/%Y")
            ws.cell(row=fila_destino, column=6).value = st.session_state.cat_temp
            ws.cell(row=fila_destino, column=7).value = st.session_state.detalles_temp.upper()

            wb.save(archivo)
            
            # Guardar en memoria para el botón de deshacer
            st.session_state.ultimo_guardado = {
                "fila": fila_destino,
                "mes": mes_p,
                "concepto": st.session_state.concepto_temp,
                "importe": st.session_state.importe_temp
            }
            
            st.success(f"✅ ¡Guardado en la fila {fila_destino}!")
            st.balloons()
            time.sleep(1)
            finalizar_y_limpiar()
            
        except ValueError as ve:
            st.error(str(ve))
            if st.button("Volver"):
                st.session_state.paso = "formulario"
                st.rerun()
        except PermissionError:
            st.error("¡Cierra el Excel! No puedo guardar si está abierto.")
        except Exception as e:
            st.error(f"Error inesperado: {e}")

    if c_no.button("❌ CANCELAR Y CORREGIR", use_container_width=True):
        st.session_state.paso = "formulario"
        st.rerun()